"""Student management routes"""
import uuid
import io
import shutil
from pathlib import Path
from fastapi import APIRouter, HTTPException, Request, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import datetime, timezone

from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

import sys; sys.path.insert(0, "/app/backend"); from config import db, logger, STUDENT_DOCUMENTS_PATH
from models.students import StudentCreate, StudentUpdate, StudentResponse, ConvertLeadToStudent
from utils.auth import get_current_user, require_roles
from utils.helpers import create_audit_log

router = APIRouter(prefix="/students", tags=["students"])


def generate_institutional_email(full_name: str) -> str:
    """Generate an institutional email from student name"""
    parts = full_name.lower().strip().split()
    if len(parts) >= 2:
        email_base = f"{parts[0]}.{parts[-1]}"
    else:
        email_base = parts[0] if parts else "estudiante"
    
    # Remove accents and special characters
    import unicodedata
    email_base = unicodedata.normalize('NFKD', email_base).encode('ASCII', 'ignore').decode()
    email_base = ''.join(c for c in email_base if c.isalnum() or c == '.')
    
    return f"{email_base}@ucic.edu.mx"


@router.post("", response_model=StudentResponse)
async def create_student(student_data: StudentCreate, request: Request):
    """Create a new student"""
    await require_roles(["admin", "gerente"])(request)
    
    student_id = f"student_{uuid.uuid4().hex[:12]}"
    now = datetime.now(timezone.utc)
    
    # Generate institutional email if not provided
    institutional_email = student_data.institutional_email
    if not institutional_email:
        institutional_email = generate_institutional_email(student_data.full_name)
        # Check for duplicates and add number if needed
        base_email = institutional_email.replace("@ucic.edu.mx", "")
        counter = 1
        while await db.students.find_one({"institutional_email": institutional_email}, {"_id": 0}):
            institutional_email = f"{base_email}{counter}@ucic.edu.mx"
            counter += 1
    
    # Create document folder for student
    student_folder = STUDENT_DOCUMENTS_PATH / student_id
    student_folder.mkdir(exist_ok=True)
    
    student = {
        "student_id": student_id,
        "full_name": student_data.full_name,
        "email": student_data.email,
        "phone": student_data.phone,
        "career_id": student_data.career_id,
        "career_name": student_data.career_name,
        "institutional_email": institutional_email,
        "lead_id": student_data.lead_id,
        "documents": [],
        "attendance": [],
        "custom_fields": {},
        "is_active": True,
        "created_at": now.isoformat(),
        "updated_at": now.isoformat()
    }
    
    await db.students.insert_one(student)
    student.pop("_id", None)
    
    logger.info(f"Student created: {student_id}")
    return StudentResponse(**student)


# Custom fields management
@router.get("/custom-fields")
async def get_custom_fields(request: Request):
    """Get all custom field definitions"""
    await get_current_user(request)
    
    fields = await db.custom_fields.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    return {"fields": fields}


@router.post("/custom-fields")
async def create_custom_field(request: Request):
    """Create a custom field definition"""
    current_user = await require_roles(["admin", "gerente"])(request)
    body = await request.json()
    
    field_id = f"field_{uuid.uuid4().hex[:8]}"
    now = datetime.now(timezone.utc)
    
    # Get max order
    last_field = await db.custom_fields.find_one(sort=[("order", -1)])
    next_order = (last_field["order"] + 1) if last_field else 0
    
    field = {
        "field_id": field_id,
        "field_name": body.get("field_name", ""),
        "field_type": body.get("field_type", "text"),
        "options": body.get("options", []),
        "required": body.get("required", False),
        "visible_to_students": body.get("visible_to_students", True),
        "editable_by_supervisor": body.get("editable_by_supervisor", True),
        "order": next_order,
        "created_at": now.isoformat(),
        "created_by": current_user["user_id"]
    }
    
    await db.custom_fields.insert_one(field)
    field.pop("_id", None)
    
    await create_audit_log(
        entity_type="custom_field",
        entity_id=field_id,
        action="create",
        performed_by=current_user,
        new_value=field,
        request=request
    )
    
    return {"field": field, "message": "Campo creado exitosamente"}


@router.put("/custom-fields/{field_id}")
async def update_custom_field(field_id: str, request: Request):
    """Update a custom field definition"""
    current_user = await require_roles(["admin", "gerente"])(request)
    body = await request.json()
    
    field = await db.custom_fields.find_one({"field_id": field_id}, {"_id": 0})
    if not field:
        raise HTTPException(status_code=404, detail="Campo no encontrado")
    
    update_data = {k: v for k, v in body.items() if k in ["field_name", "field_type", "options", "required", "visible_to_students", "editable_by_supervisor", "order"]}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.custom_fields.update_one({"field_id": field_id}, {"$set": update_data})
    
    await create_audit_log(
        entity_type="custom_field",
        entity_id=field_id,
        action="update",
        performed_by=current_user,
        old_value=field,
        new_value=update_data,
        request=request
    )
    
    return {"message": "Campo actualizado"}


@router.delete("/custom-fields/{field_id}")
async def delete_custom_field(field_id: str, request: Request):
    """Delete a custom field definition"""
    current_user = await require_roles(["admin"])(request)
    
    result = await db.custom_fields.delete_one({"field_id": field_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Campo no encontrado")
    
    # Remove field values from all students
    await db.students.update_many({}, {"$unset": {f"custom_fields.{field_id}": ""}})
    
    await create_audit_log(
        entity_type="custom_field",
        entity_id=field_id,
        action="delete",
        performed_by=current_user,
        request=request
    )
    
    return {"message": "Campo eliminado"}


# Change requests (approval workflow)
@router.get("/change-requests")
async def get_change_requests(request: Request, status: Optional[str] = None):
    """Get all change requests (for approval)"""
    await require_roles(["admin", "gerente"])(request)
    
    query = {}
    if status:
        query["status"] = status
    
    requests_list = await db.change_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return {"requests": requests_list}


@router.post("/change-requests/{request_id}/approve")
async def approve_change_request(request_id: str, request: Request):
    """Approve a change request"""
    current_user = await require_roles(["admin", "gerente"])(request)
    
    change_req = await db.change_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not change_req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    if change_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="La solicitud ya fue procesada")
    
    now = datetime.now(timezone.utc)
    
    # Apply the change
    student = await db.students.find_one({"student_id": change_req["student_id"]}, {"_id": 0})
    if student:
        custom_fields = student.get("custom_fields", {})
        custom_fields[change_req["field_id"]] = change_req["new_value"]
        
        await db.students.update_one(
            {"student_id": change_req["student_id"]},
            {"$set": {"custom_fields": custom_fields, "updated_at": now.isoformat()}}
        )
    
    # Update request status
    await db.change_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "approved",
            "approved_by_id": current_user["user_id"],
            "approved_by_name": current_user["name"],
            "resolved_at": now.isoformat()
        }}
    )
    
    # Create audit log
    await create_audit_log(
        entity_type="student",
        entity_id=change_req["student_id"],
        action="approve",
        field_changed=change_req["field_name"],
        old_value=change_req["old_value"],
        new_value=change_req["new_value"],
        performed_by={"user_id": change_req["requested_by_id"], "name": change_req["requested_by_name"], "role": "supervisor"},
        authorized_by=current_user,
        request=request
    )
    
    return {"message": "Cambio aprobado"}


@router.post("/change-requests/{request_id}/reject")
async def reject_change_request(request_id: str, request: Request):
    """Reject a change request"""
    current_user = await require_roles(["admin", "gerente"])(request)
    
    change_req = await db.change_requests.find_one({"request_id": request_id}, {"_id": 0})
    if not change_req:
        raise HTTPException(status_code=404, detail="Solicitud no encontrada")
    
    if change_req["status"] != "pending":
        raise HTTPException(status_code=400, detail="La solicitud ya fue procesada")
    
    now = datetime.now(timezone.utc)
    
    await db.change_requests.update_one(
        {"request_id": request_id},
        {"$set": {
            "status": "rejected",
            "approved_by_id": current_user["user_id"],
            "approved_by_name": current_user["name"],
            "resolved_at": now.isoformat()
        }}
    )
    
    await create_audit_log(
        entity_type="change_request",
        entity_id=request_id,
        action="reject",
        performed_by=current_user,
        request=request
    )
    
    return {"message": "Cambio rechazado"}


# Audit logs
@router.get("/audit-logs")
async def get_audit_logs(request: Request, entity_type: Optional[str] = None, entity_id: Optional[str] = None):
    """Get audit logs"""
    await require_roles(["admin", "gerente"])(request)
    
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(1000)
    return {"logs": logs}


# Export endpoints
@router.get("/export/excel")
async def export_students_excel(request: Request):
    """Export students to Excel"""
    await require_roles(["admin", "gerente"])(request)
    
    students = await db.students.find({}, {"_id": 0}).to_list(10000)
    custom_fields = await db.custom_fields.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"
    
    # Headers
    headers = ["ID", "Nombre", "Email", "Teléfono", "Carrera", "Email Institucional", "Fecha Inscripción"]
    for field in custom_fields:
        headers.append(field["field_name"])
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = cell.font.copy(bold=True)
    
    # Data rows
    for row_num, student in enumerate(students, 2):
        ws.cell(row=row_num, column=1, value=student.get("student_id", ""))
        ws.cell(row=row_num, column=2, value=student.get("full_name", ""))
        ws.cell(row=row_num, column=3, value=student.get("email", ""))
        ws.cell(row=row_num, column=4, value=student.get("phone", ""))
        ws.cell(row=row_num, column=5, value=student.get("career_name", ""))
        ws.cell(row=row_num, column=6, value=student.get("institutional_email", ""))
        ws.cell(row=row_num, column=7, value=student.get("created_at", "")[:10] if student.get("created_at") else "")
        
        custom_values = student.get("custom_fields", {})
        for col_offset, field in enumerate(custom_fields, 8):
            value = custom_values.get(field["field_id"], "")
            ws.cell(row=row_num, column=col_offset, value=str(value) if value is not None else "")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"estudiantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/export/pdf")
async def export_students_pdf(request: Request):
    """Export students to PDF"""
    await require_roles(["admin", "gerente"])(request)
    
    students = await db.students.find({}, {"_id": 0}).to_list(10000)
    custom_fields = await db.custom_fields.find({}, {"_id": 0}).sort("order", 1).to_list(100)
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(letter), topMargin=0.5*inch, bottomMargin=0.5*inch)
    
    elements = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=1
    )
    elements.append(Paragraph("UCIC - Lista de Estudiantes", title_style))
    elements.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Table headers
    headers = ["Nombre", "Email", "Teléfono", "Carrera", "Email Inst."]
    # Limit custom fields for PDF to avoid overflow
    for field in custom_fields[:3]:
        headers.append(field["field_name"][:15])
    
    # Table data
    data = [headers]
    for student in students:
        row = [
            student.get("full_name", "")[:25],
            student.get("email", "")[:25],
            student.get("phone", "")[:15],
            student.get("career_name", "")[:20],
            (student.get("institutional_email", "") or "")[:20]
        ]
        
        custom_values = student.get("custom_fields", {})
        for field in custom_fields[:3]:
            val = str(custom_values.get(field["field_id"], ""))[:15]
            row.append(val)
        
        data.append(row)
    
    # Create table
    col_widths = [1.5*inch, 1.8*inch, 1.2*inch, 1.5*inch, 1.5*inch]
    col_widths.extend([1*inch] * min(len(custom_fields), 3))
    
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e293b')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')])
    ]))
    
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 20))
    elements.append(Paragraph(f"Total: {len(students)} estudiantes", styles['Normal']))
    
    doc.build(elements)
    output.seek(0)
    
    filename = f"estudiantes_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("", response_model=List[StudentResponse])
async def get_students(request: Request):
    """Get all students"""
    await get_current_user(request)
    
    students = await db.students.find({}, {"_id": 0}).to_list(1000)
    return [StudentResponse(**s) for s in students]


@router.get("/{student_id}", response_model=StudentResponse)
async def get_student(student_id: str, request: Request):
    """Get a single student"""
    await get_current_user(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    return StudentResponse(**student)


@router.put("/{student_id}", response_model=StudentResponse)
async def update_student(student_id: str, student_data: StudentUpdate, request: Request):
    """Update a student"""
    await require_roles(["admin", "gerente"])(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    update_data = {k: v for k, v in student_data.model_dump().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.students.update_one({"student_id": student_id}, {"$set": update_data})
    
    updated_student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    return StudentResponse(**updated_student)


@router.delete("/{student_id}")
async def delete_student(student_id: str, request: Request):
    """Delete a student and their documents"""
    await require_roles(["admin"])(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Delete document folder
    student_folder = STUDENT_DOCUMENTS_PATH / student_id
    if student_folder.exists():
        shutil.rmtree(student_folder)
    
    await db.students.delete_one({"student_id": student_id})
    
    return {"message": "Estudiante eliminado"}


# Document management
@router.post("/{student_id}/documents")
async def upload_document(
    student_id: str,
    request: Request,
    file: UploadFile = File(...),
    document_type: str = "Otro"
):
    """Upload a document for a student"""
    await require_roles(["admin", "gerente", "supervisor"])(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Create student folder if not exists
    student_folder = STUDENT_DOCUMENTS_PATH / student_id
    student_folder.mkdir(exist_ok=True)
    
    # Generate unique filename
    document_id = f"doc_{uuid.uuid4().hex[:8]}"
    file_extension = Path(file.filename).suffix if file.filename else ".pdf"
    safe_filename = f"{document_id}{file_extension}"
    file_path = student_folder / safe_filename
    
    # Save file
    with open(file_path, "wb") as buffer:
        content = await file.read()
        buffer.write(content)
    
    # Add document reference to student
    document = {
        "document_id": document_id,
        "name": document_type,
        "filename": safe_filename,
        "original_filename": file.filename,
        "uploaded_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.students.update_one(
        {"student_id": student_id},
        {"$push": {"documents": document}}
    )
    
    return {"message": "Documento subido exitosamente", "document": document}


@router.delete("/{student_id}/documents/{document_id}")
async def delete_document(student_id: str, document_id: str, request: Request):
    """Delete a document"""
    await require_roles(["admin", "gerente"])(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Find document
    document = None
    for doc in student.get("documents", []):
        if doc["document_id"] == document_id:
            document = doc
            break
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    # Delete file
    file_path = STUDENT_DOCUMENTS_PATH / student_id / document["filename"]
    if file_path.exists():
        file_path.unlink()
    
    # Remove from database
    await db.students.update_one(
        {"student_id": student_id},
        {"$pull": {"documents": {"document_id": document_id}}}
    )
    
    return {"message": "Documento eliminado"}


@router.get("/{student_id}/documents/{document_id}/download")
async def download_document(student_id: str, document_id: str, request: Request):
    """Download a document"""
    from fastapi.responses import FileResponse
    
    await get_current_user(request)
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Find document
    document = None
    for doc in student.get("documents", []):
        if doc["document_id"] == document_id:
            document = doc
            break
    
    if not document:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    
    file_path = STUDENT_DOCUMENTS_PATH / student_id / document["filename"]
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    
    # Determine content type based on file extension
    extension = document["filename"].lower().split(".")[-1] if "." in document["filename"] else ""
    content_types = {
        "pdf": "application/pdf",
        "doc": "application/msword",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "gif": "image/gif"
    }
    media_type = content_types.get(extension, "application/octet-stream")
    
    original_filename = document.get("original_filename", document["filename"])
    
    return FileResponse(
        path=str(file_path),
        media_type=media_type,
        filename=original_filename
    )


# Attendance management
@router.post("/{student_id}/attendance")
async def record_attendance(student_id: str, request: Request):
    """Record attendance for a student"""
    await require_roles(["admin", "gerente", "supervisor", "maestro"])(request)
    
    body = await request.json()
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    attendance_record = {
        "date": body.get("date", datetime.now(timezone.utc).strftime("%Y-%m-%d")),
        "subject": body.get("subject", ""),
        "teacher_id": body.get("teacher_id"),
        "teacher_name": body.get("teacher_name"),
        "status": body.get("status", "presente"),
        "notes": body.get("notes")
    }
    
    await db.students.update_one(
        {"student_id": student_id},
        {"$push": {"attendance": attendance_record}}
    )
    
    return {"message": "Asistencia registrada"}


@router.put("/{student_id}/custom-fields")
async def update_student_custom_fields(student_id: str, request: Request):
    """Update custom field values for a student"""
    current_user = await get_current_user(request)
    body = await request.json()
    
    student = await db.students.find_one({"student_id": student_id}, {"_id": 0})
    if not student:
        raise HTTPException(status_code=404, detail="Estudiante no encontrado")
    
    # Check permissions
    user_role = current_user["role"]
    
    # Students can only view
    if user_role == "alumno":
        raise HTTPException(status_code=403, detail="Los alumnos no pueden modificar datos")
    
    # Supervisors need approval for changes
    if user_role == "supervisor":
        changes = body.get("fields", {})
        old_custom_fields = student.get("custom_fields", {})
        
        for field_id, new_value in changes.items():
            field_def = await db.custom_fields.find_one({"field_id": field_id}, {"_id": 0})
            if not field_def:
                continue
                
            if not field_def.get("editable_by_supervisor", True):
                continue
            
            old_value = old_custom_fields.get(field_id)
            if old_value != new_value:
                request_id = f"req_{uuid.uuid4().hex[:8]}"
                now = datetime.now(timezone.utc)
                
                change_request = {
                    "request_id": request_id,
                    "student_id": student_id,
                    "student_name": student["full_name"],
                    "field_id": field_id,
                    "field_name": field_def["field_name"],
                    "old_value": old_value,
                    "new_value": new_value,
                    "requested_by_id": current_user["user_id"],
                    "requested_by_name": current_user["name"],
                    "status": "pending",
                    "created_at": now.isoformat()
                }
                
                await db.change_requests.insert_one(change_request)
        
        return {"message": "Solicitud de cambio enviada para aprobación", "requires_approval": True}
    
    # Admin/Gerente can update directly
    changes = body.get("fields", {})
    old_custom_fields = student.get("custom_fields", {})
    
    for field_id, new_value in changes.items():
        field_def = await db.custom_fields.find_one({"field_id": field_id}, {"_id": 0})
        if field_def:
            old_value = old_custom_fields.get(field_id)
            
            await create_audit_log(
                entity_type="student",
                entity_id=student_id,
                action="update",
                field_changed=field_def["field_name"],
                old_value=old_value,
                new_value=new_value,
                performed_by=current_user,
                request=request
            )
    
    await db.students.update_one(
        {"student_id": student_id},
        {"$set": {"custom_fields": changes, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    return {"message": "Campos actualizados"}


